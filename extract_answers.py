import openpyxl
from bs4 import BeautifulSoup
import re
import tkinter as tk
from tkinter import filedialog
import sys
from difflib import SequenceMatcher

def clean_text(text):
    """
    【升级版数据清洗】：去除所有空白字符、标点符号、括号等。
    只保留汉字、英文字母和数字。
    彻底解决（）和()、全角半角标点带来的匹配失败问题。
    """
    if text is None or text == "":
        return ""
    text = str(text).lower()
    # \w 匹配字母数字下划线，\u4e00-\u9fa5 匹配汉字
    # 这里将所有不在这个范围内的字符（如括号、问号、冒号、空格）全部替换为空
    text = re.sub(r'[^\w\u4e00-\u9fa5]', '', text)
    # 将填空题中常见的下划线也去掉
    return text.replace('_', '')

def col_letter_to_index(col_letter):
    col_letter = col_letter.strip().upper()
    index = 0
    for char in col_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1

def parse_opt_cols(opt_cols_str):
    indices = []
    parts = opt_cols_str.split(',')
    for part in parts:
        part = part.strip().upper()
        if not part:
            continue
        if '-' in part:
            start_letter, end_letter = part.split('-')
            start_idx = col_letter_to_index(start_letter.strip())
            end_idx = col_letter_to_index(end_letter.strip())
            for i in range(start_idx, end_idx + 1):
                indices.append(i)
        else:
            indices.append(col_letter_to_index(part))
    return sorted(list(set(indices)))

def safe_get_cell(row, index):
    """适配 openpyxl 元组的获取方式"""
    if index < len(row):
        val = row[index]
        return val if val is not None else ""
    return ""

def extract_answers_from_excel(excel_path, q_col, ans_col, opt_cols):
    excel_dict = {}
    
    q_idx = col_letter_to_index(q_col)
    ans_idx = col_letter_to_index(ans_col)
    opt_indices = parse_opt_cols(opt_cols)
    
    true_keywords = ['正确', '对', '√', 'TRUE', 'Ture']
    false_keywords = ['错误', '错', '×', 'FALSE', 'False']
    
    # === 改用 openpyxl 直接按行读取 ===
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # iter_rows 每次返回一行数据的元组
    for row in sheet.iter_rows(values_only=True):
        question = clean_text(safe_get_cell(row, q_idx))
        if not question:
            continue
            
        raw_answer = str(safe_get_cell(row, ans_idx)).strip().upper()
        
        if any(kw == raw_answer for kw in true_keywords):
            data = {'type': 'judge', 'value': '✅'}
        elif any(kw == raw_answer for kw in false_keywords):
            data = {'type': 'judge', 'value': '❌'}
        else:
            correct_texts = []
            for letter in raw_answer:
                if 'A' <= letter <= 'Z':
                    opt_offset = ord(letter) - ord('A')
                    if opt_offset < len(opt_indices):
                        opt_val = safe_get_cell(row, opt_indices[opt_offset])
                        correct_texts.append(clean_text(opt_val))
            data = {'type': 'choice', 'value': correct_texts}
            
        if question not in excel_dict:
            excel_dict[question] = []
        excel_dict[question].append(data)
            
    return excel_dict

def get_best_match_from_fuzzy(clean_q, excel_dict, expected_type):
    """
    【升级版模糊匹配】：增加题型感知机制
    如果网页是选择题，它只会去 Excel 里寻找答案结构是选择题的题目，绝不会跨界匹配判断题。
    """
    best_match = None
    highest_ratio = 0.0
    
    # 优先在题型匹配的题目中寻找相似度最高的
    for ex_q, candidates in excel_dict.items():
        if not any(cand['type'] == expected_type for cand in candidates):
            continue
            
        ratio = SequenceMatcher(None, clean_q, ex_q).ratio()
        if ratio > highest_ratio:
            highest_ratio = ratio
            best_match = ex_q
            
    # 极端容错：如果符合题型的全部没找到，再降级去全库搜一次
    if best_match is None or highest_ratio < 0.2:
        for ex_q in excel_dict.keys():
            ratio = SequenceMatcher(None, clean_q, ex_q).ratio()
            if ratio > highest_ratio:
                highest_ratio = ratio
                best_match = ex_q
                
    return best_match, highest_ratio

def parse_html_and_match(html_path, excel_dict):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
        
    questions_divs = soup.find_all('div', class_='questions')
    
    for q in questions_divs:
        title_div = q.find('div', class_='title')
        if not title_div: 
            continue
        
        spans = title_div.find_all('span')
        if len(spans) < 2: 
            continue
            
        # 提取网页上注明的题型（单选/多选/判断）
        em_tag = title_div.find('em')
        q_type_str = em_tag.get_text(strip=True) if em_tag else ""
        expected_type = 'judge' if '判断' in q_type_str else 'choice'
        
        q_num = spans[0].get_text(strip=True)
        q_text = spans[1].get_text(strip=True)
        clean_q = clean_text(q_text)
        
        html_options = {}
        lis = q.find_all('li')
        for li in lis:
            em = li.find('em')
            span = li.find('span')
            if em and span:
                letter = em.get_text(strip=True)
                opt_text = clean_text(span.get_text(strip=True))
                if letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    html_options[opt_text] = letter
                    
        print(f"题号: {q_num} {q_text}")
        
        candidates = []
        
        if clean_q in excel_dict:
            candidates = excel_dict[clean_q]
        else:
            # 传入预期题型，防窜台
            best_match_q, _ = get_best_match_from_fuzzy(clean_q, excel_dict, expected_type)
            if best_match_q:
                candidates = excel_dict[best_match_q]
                print(f"  [!] 未找到精确题干，已智能匹配最相似同类型题目：\n  -> {best_match_q[:25]}...")
                
        if not candidates:
            print("正确答案: [题库为空或无法匹配]")
            print("-" * 40)
            continue
            
        best_candidate = None
        max_hit_score = -1
        
        for cand in candidates:
            if cand['type'] == 'judge':
                hit_score = 0
            else:
                correct_texts = cand['value']
                hit_score = sum(1 for ct in correct_texts if ct in html_options)
            
            # 如果存在多道同名题，优先取类型一致的，其次取选项匹配度最高的
            if cand['type'] == expected_type:
                hit_score += 100 # 给题型一致的候选答案加个绝对高分权重
                
            if hit_score > max_hit_score:
                max_hit_score = hit_score
                best_candidate = cand

        if best_candidate['type'] == 'judge':
            print(f"正确答案: {best_candidate['value']}")
        else:
            correct_texts = best_candidate['value']
            final_answers = []
            
            for ct in correct_texts:
                if ct in html_options:
                    final_answers.append((html_options[ct], ct))
                else:
                    final_answers.append(('?', f"未找到文本匹配: {ct[:10]}..."))
            
            final_answers.sort(key=lambda x: x[0])
            letters_str = "".join([x[0] for x in final_answers])
            texts_str = ", ".join([x[1] for x in final_answers])
            
            print(f"正确答案: {letters_str} ({texts_str})")
            
        print("-" * 40)

def main():
    root = tk.Tk()
    root.withdraw()

    print("=== 试卷答案提取工具 ===\n")
    
    print("【第一步：配置 Excel 数据列】")
    q_col = input("请输入【题干】所在列（例如 A）: ").strip()
    ans_col = input("请输入【答案代号】所在列（例如 H）: ").strip()
    opt_cols = input("请输入【选项】所在列（支持 B-E 或 B,C,D）: ").strip()
    
    if not (q_col and ans_col and opt_cols):
        print("\n输入不完整，程序退出。")
        sys.exit()

    print("\n【第二步：选择文件】")
    print("请在弹出的窗口中选择包含答案的 Excel 文件...")
    excel_file = filedialog.askopenfilename(
        title="选择 Excel 文件",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not excel_file:
        sys.exit()

    print("请在弹出的窗口中选择要解析的 HTML 文件...")
    html_file = filedialog.askopenfilename(
        title="选择 HTML 文件",
        filetypes=[("HTML files", "*.html *.htm")]
    )
    if not html_file:
        sys.exit()

    print("\n>>> 开始解析...\n")
    try:
        excel_data = extract_answers_from_excel(excel_file, q_col, ans_col, opt_cols)
        parse_html_and_match(html_file, excel_data)
        print("\n=== 所有题目处理完成 ===")
    except Exception as e:
        print(f"\n[错误] 运行过程中出现问题: {e}")
        
    input("\n按回车键退出程序...")

if __name__ == "__main__":
    main()